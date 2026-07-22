#!/usr/bin/env python3
"""Extract block/GP-level partner presence (count-only) from the source xlsx files
and merge a `blockcov` field into each district in model.json.

Presence sources with block-level resolution: Common Ground + TRI (via TRI sheet
and SOTH village list, org=TRIF). The other 12 landscape partners are mapped at
district level only, so they never appear here. This is deliberately labelled in
the UI as "known block-level presence", not total coverage.
"""
import json, re
import openpyxl

# --- district-name normalisation to the 24-name canon ---
CANON = json.load(open("model.json"))["canon"]
_norm = {c.lower().replace("-", " ").replace(".", "").strip(): c for c in CANON}
ALIAS = {
    "saraikela kharswan": "Saraikela-Kharsawan",
    "seraikela kharsawan": "Saraikela-Kharsawan",
    "west singbhum": "West Singhbhum",
    "giridh": "Giridih",
    "sahibganj": "Sahibganj",
    "purbi singhbhum": "East Singhbhum",
    "pashchimi singhbhum": "West Singhbhum",
}
def canon(name):
    if not name: return None
    k = str(name).lower().replace("-", " ").replace(".", "").strip()
    k = re.sub(r"\s+", " ", k)
    if k in ALIAS: return ALIAS[k]
    if k in _norm: return _norm[k]
    return None  # out-of-state / unmatched -> dropped

def clean(s):
    return re.sub(r"\s+", " ", str(s)).strip() if s else ""

# district -> { blockkey : {"name":..,"by":set(),"villages":set()} }
from collections import defaultdict
cov = defaultdict(dict)

def add(dist, block, source, village=None):
    d = canon(dist); b = clean(block)
    if not d or not b: return
    key = b.lower()
    e = cov[d].setdefault(key, {"name": b, "by": set(), "villages": set()})
    e["by"].add(source)
    if village: e["villages"].add(clean(village))

# --- 1. Common Ground (Jharkhand sheet: District | Block(comma list) | No of Blocks) ---
wb = openpyxl.load_workbook("Common Ground - List of blocks (1).xlsx", read_only=True, data_only=True)
for r in list(wb["Jharkhand"].iter_rows(values_only=True))[4:]:
    dist, blocks = r[2], r[3]
    if dist and blocks:
        for b in str(blocks).split(","):
            add(dist, b, "Common Ground")

# --- 2. TRI (Geographic Presence: State | District | Blocks(Community Action Lab)) ---
wb = openpyxl.load_workbook("TRI Geographical Presence - jul 2026.xlsx", read_only=True, data_only=True)
cur_state = None
for r in list(wb["Geographic Presence"].iter_rows(values_only=True))[4:]:
    if r[2]: cur_state = clean(r[2])
    if cur_state == "Jharkhand" and r[3] and r[4]:
        for b in str(r[4]).split(","):
            add(r[3], b, "TRI")

# --- 3. SOTH (Place list: Organisation | Village | Block/GP | District | State), Jharkhand only ---
wb = openpyxl.load_workbook("SOTH places list.xlsx", read_only=True, data_only=True)
for r in list(wb["Place list"].iter_rows(values_only=True))[2:]:
    if not r or r[4] != "Jharkhand": continue
    org, village, blockgp, dist = r[0], r[1], r[2], r[3]
    src = "TRI" if org and "trif" in str(org).lower() else clean(org)
    add(dist, blockgp, src, village)

# --- fuzzy de-dupe near-identical block names within a district (spelling variants) ---
def lev(a, b):
    a, b = a.lower(), b.lower()
    if a == b: return 0
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + (ca != cb)))
        prev = cur
    return prev[-1]

for d, blocks in cov.items():
    items = list(blocks.values())
    merged, used = [], set()
    for i, e in enumerate(items):
        if i in used: continue
        for j in range(i + 1, len(items)):
            if j in used: continue
            f = items[j]
            if len(e["name"]) >= 5 and len(f["name"]) >= 5 and lev(e["name"], f["name"]) <= 2:
                # keep the variant with more evidence (villages, then longer name)
                keep, drop = (e, f) if (len(e["villages"]), len(e["name"])) >= (len(f["villages"]), len(f["name"])) else (f, e)
                keep["by"] |= drop["by"]; keep["villages"] |= drop["villages"]
                e = keep; used.add(j)
        merged.append(e)
    cov[d] = {e["name"].lower(): e for e in merged}

# --- merge into model.json ---
model = json.load(open("model.json"))
tot_blocks = tot_dist = 0
for d in CANON:
    entries = sorted(cov.get(d, {}).values(), key=lambda e: e["name"].lower())
    if entries:
        tot_dist += 1; tot_blocks += len(entries)
    model["districts"][d]["blockcov"] = [
        {"name": e["name"], "by": sorted(e["by"]),
         "villages": sorted(v for v in e["villages"] if v)}
        for e in entries
    ]

json.dump(model, open("model.json", "w"), ensure_ascii=False, separators=(",", ":"))
print(f"blockcov merged: {tot_blocks} block/GPs with known presence across {tot_dist} districts")
for d in CANON:
    bc = model["districts"][d]["blockcov"]
    if bc:
        srcs = sorted({s for e in bc for s in e["by"]})
        print(f"  {d}: {len(bc)} block/GP  ({', '.join(srcs)})")
